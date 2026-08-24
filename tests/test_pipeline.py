import json

from openpyxl import load_workbook

from fincompiler.exception_workflow import read_exception_workflow, update_exception_item
from fincompiler.pipeline import compile_pack
from fincompiler.reporting import write_management_pack_excel


def test_end_to_end_pack_is_blocked_by_reconciliation(demo_dir, tmp_path):
    report = compile_pack(demo_dir, tmp_path / "out", tmp_path / "mappings.json")
    assert report["output_readiness"] == "BLOCKED"
    assert report["reconciliation"]["variance"] == "2706.00"
    saved = json.loads((tmp_path / "out" / "management_pack.json").read_text())
    assert saved["pvm"]["totals"]["bridge_check"] == "0.00"
    assert saved["management_pack_excel"] == "management_pack.xlsx"
    exception_state = json.loads((tmp_path / "out" / "exception_workflow.json").read_text(encoding="utf-8"))
    assert exception_state["active_count"] == len(saved["exceptions"])
    workbook = load_workbook(tmp_path / "out" / "management_pack.xlsx", data_only=True, read_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Control Checks",
        "Action Plan",
        "Exceptions",
        "Sales vs GL",
        "Budget vs Actual",
        "FX Evidence",
        "FX Sources",
        "Lineage",
        "Source Audit",
        "Mappings",
    ]
    assert workbook["Summary"]["B3"].value == "BLOCKED"
    assert workbook["Sales vs GL"]["B5"].value == 2706
    assert workbook["Control Checks"]["F6"].value == "FAIL"
    assert workbook["Lineage"].max_row > 3
    workbook.close()
    first_item = exception_state["items"][0]
    update_exception_item(
        tmp_path / "out",
        first_item["exception_id"],
        status="INVESTIGATING",
        owner="Finance Manager",
        note="Checking the source export",
    )
    write_management_pack_excel(tmp_path / "out", report, read_exception_workflow(tmp_path / "out"))
    refreshed = load_workbook(tmp_path / "out" / "management_pack.xlsx", data_only=True, read_only=True)
    assert refreshed["Exceptions"]["C4"].value == "INVESTIGATING"
    assert refreshed["Exceptions"]["D4"].value == "Finance Manager"
    refreshed.close()
    rendered = (tmp_path / "out" / "management_pack.html").read_text(encoding="utf-8")
    assert "FinCompiler Management Pack" in rendered
    assert "2706.00" in rendered
    assert saved["run_manifest"]["run_id"] in rendered

import json

from openpyxl import load_workbook

from fincompiler.commerce import compile_commerce_pack


def test_cross_border_pack_matches_platform_cash_and_calculates_sku_profit(tmp_path):
    report = compile_commerce_pack(
        "demo/cross_border_seller",
        tmp_path / "out",
        tmp_path / "commerce-memory.json",
    )

    assert report["output_readiness"] == "READY"
    assert [item["status"] for item in report["payout_reconciliation"]] == ["PASS", "PASS"]
    profit = {(item["platform"], item["sku"]): item for item in report["sku_profitability"]}
    assert profit[("Amazon", "AMZ-A")]["profit"] == "430.00"
    assert profit[("Shopify", "SHP-B")]["profit"] == "93.00"
    assert report["exceptions"] == []
    saved = json.loads((tmp_path / "out" / "commerce_pack.json").read_text(encoding="utf-8"))
    assert saved["calculation_policy"].startswith("All amounts are calculated with Decimal")
    assert saved["source_lineage"]["amazon_settlements"][0]["file"].endswith("amazon_settlements.csv")
    workbook = load_workbook(tmp_path / "out" / "commerce_pack.xlsx", data_only=True, read_only=True)
    assert workbook.sheetnames == ["Summary", "Payout Reconciliation", "SKU Profitability", "Exceptions", "Source Lineage"]
    assert workbook["Summary"]["B3"].value == "READY"
    workbook.close()


def test_cross_border_pack_identifies_specific_payout_bank_variance(tmp_path):
    source = tmp_path / "input"
    source.mkdir()
    for path in __import__("pathlib").Path("demo/cross_border_seller").iterdir():
        if path.is_file():
            __import__("shutil").copy2(path, source / path.name)
    bank = source / "bank.csv"
    bank.write_text(bank.read_text(encoding="utf-8").replace("980.00,USD,Amazon", "975.00,USD,Amazon"), encoding="utf-8")

    report = compile_commerce_pack(source, tmp_path / "out", tmp_path / "memory.json")

    assert report["output_readiness"] == "BLOCKED"
    amazon = next(item for item in report["payout_reconciliation"] if item["platform"] == "Amazon")
    assert amazon["bank_record_id"] == "BANK-001"
    assert amazon["difference"] == "5.00"
    assert amazon["reason"] == "BANK_OR_PROCESSOR_DEDUCTION"
    issue = next(item for item in report["exceptions"] if item["code"] == "PAYOUT_BANK_VARIANCE")
    assert issue["context"]["payout_id"] == "amazon:AMZ-ST-1001"


def test_cross_border_pack_applies_approved_dated_fx_to_platform_and_costs(tmp_path):
    from pathlib import Path
    import shutil

    source = tmp_path / "fx-input"
    shutil.copytree(Path("demo/cross_border_seller"), source)
    (source / "fx_rates.csv").write_text(
        "effective_date,base_currency,quote_currency,rate,rate_type,provider,source_url,fetched_at,raw_sha256\n"
        "2026-01-01,USD,AED,3.6725,transaction,Company Treasury,https://example.invalid/rates,2026-01-01T00:00:00Z,\n"
        "2026-07-01,USD,AED,3.6725,transaction,Company Treasury,https://example.invalid/rates,2026-07-01T00:00:00Z,\n",
        encoding="utf-8",
    )
    (source / "commerce_config.json").write_text(json.dumps({
        "business_name": "FX seller",
        "base_currency": "AED",
        "fx_rate_book": "fx_rates.csv",
        "bank_match_window_days": 7,
        "fx_policy": {"rate_type": "transaction", "max_lookback_days": 31, "allow_inverse": True, "allow_cross": True, "triangulation_currency": "EUR"},
    }), encoding="utf-8")
    bank = source / "bank.csv"
    bank.write_text(
        bank.read_text(encoding="utf-8")
        .replace("980.00,USD,Amazon", "3599.05,AED,Amazon")
        .replace("233.00,USD,Shopify", "855.69,AED,Shopify"),
        encoding="utf-8",
    )

    report = compile_commerce_pack(source, tmp_path / "out", tmp_path / "memory.json")

    assert report["output_readiness"] == "READY"
    assert len(report["fx"]["applications"]) > 10
    assert {item["currency"] for item in report["payout_reconciliation"]} == {"AED"}
    assert [item["difference"] for item in report["payout_reconciliation"]] == ["0.00", "0.00"]


def test_cross_border_pack_blocks_true_profit_when_sku_cost_file_is_missing(tmp_path):
    from pathlib import Path
    import shutil

    source = tmp_path / "no-cost-input"
    shutil.copytree(Path("demo/cross_border_seller"), source)
    (source / "sku_costs.csv").unlink()

    report = compile_commerce_pack(source, tmp_path / "out", tmp_path / "memory.json")

    assert report["output_readiness"] == "BLOCKED"
    assert any(item["code"] == "SKU_COST_REQUIRED" for item in report["exceptions"])


def test_amazon_locale_ambiguous_amount_is_blocked_not_silently_scaled(tmp_path):
    from pathlib import Path
    import shutil

    source = tmp_path / "locale-input"
    shutil.copytree(Path("demo/cross_border_seller"), source)
    settlement = source / "amazon_settlements.csv"
    settlement.write_text(settlement.read_text(encoding="utf-8").replace("1200.00", '"1200,00"'), encoding="utf-8")

    report = compile_commerce_pack(source, tmp_path / "out", tmp_path / "memory.json")

    issue = next(item for item in report["exceptions"] if item["code"] == "TYPE_VALIDATION_FAILED")
    assert "locale-ambiguous" in issue["context"]["reason"]
    amazon = next(item for item in report["sku_profitability"] if item["platform"] == "Amazon" and item["sku"] == "AMZ-A")
    assert amazon["revenue"] != "120000.00"


def test_amazon_settlement_detail_must_add_back_to_total_amount(tmp_path):
    from pathlib import Path
    import shutil

    source = tmp_path / "incomplete-settlement"
    shutil.copytree(Path("demo/cross_border_seller"), source)
    settlement = source / "amazon_settlements.csv"
    settlement.write_text(settlement.read_text(encoding="utf-8").replace("-180.00", "-170.00"), encoding="utf-8")

    report = compile_commerce_pack(source, tmp_path / "out", tmp_path / "memory.json")

    issue = next(item for item in report["exceptions"] if item["code"] == "SETTLEMENT_ACTIVITY_INCOMPLETE")
    assert issue["context"]["settlement_id"] == "AMZ-ST-1001"
    assert issue["context"]["difference"] == "-10.00"

from __future__ import annotations

import html
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any


def _cell(value: Any) -> str:
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    return html.escape("" if value is None else str(value))


def _table(rows: list[dict[str, Any]], columns: list[tuple[str, str]]) -> str:
    if not rows:
        return '<p class="empty">No items.</p>'
    header = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body = "".join(
        "<tr>" + "".join(f"<td>{_cell(row.get(key))}</td>" for key, _ in columns) + "</tr>"
        for row in rows
    )
    return f"<div class=\"table-wrap\"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>"


def render_management_pack_html(report: dict[str, Any]) -> str:
    readiness = report["output_readiness"]
    ready_class = "ready" if readiness == "READY" else "blocked"
    workflow = report["close_workflow"]
    recon = report["reconciliation"]
    pvm = report["pvm"]
    fx = report["fx"]
    manifest = report["run_manifest"]
    tasks = workflow["tasks"]
    exceptions = report["exceptions"]
    causes = recon.get("causes", [])
    segments = pvm.get("segments", [])
    applications = fx.get("applications", [])
    sources = manifest.get("sources", [])
    return f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>FinCompiler Management Pack · {_cell(manifest['run_id'])}</title>
<style>
:root{{--navy:#14233c;--blue:#225ea8;--gold:#e4ad3a;--ink:#172033;--muted:#657086;--line:#dce2ea;--paper:#fff;--bg:#f4f6f9;--red:#a12622;--green:#176b3a}}
*{{box-sizing:border-box}} body{{margin:0;background:var(--bg);color:var(--ink);font:14px/1.5 Arial,sans-serif}}
main{{max-width:1180px;margin:0 auto;padding:32px 24px 64px}} header{{background:var(--navy);color:#fff;padding:28px;border-radius:14px}}
h1{{margin:0 0 6px;font-size:30px}} h2{{margin:30px 0 12px;font-size:20px}} .subtitle{{color:#dbe5f3}}
.badge{{display:inline-block;margin-top:16px;padding:8px 12px;border-radius:999px;font-weight:700}} .badge.ready{{background:#dff4e7;color:var(--green)}} .badge.blocked{{background:#fde5e3;color:var(--red)}}
.grid{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:16px}} .card{{background:var(--paper);border:1px solid var(--line);border-radius:10px;padding:16px}}
.label{{color:var(--muted);font-size:12px;text-transform:uppercase;letter-spacing:.04em}} .value{{font-size:22px;font-weight:700;margin-top:4px}}
.table-wrap{{overflow-x:auto;background:var(--paper);border:1px solid var(--line);border-radius:10px}} table{{width:100%;border-collapse:collapse}} th,td{{padding:10px 12px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}} th{{background:#eef2f7;color:#354056;font-size:12px}} tr:last-child td{{border-bottom:0}}
.notice{{padding:12px 14px;border-left:4px solid var(--gold);background:#fff8e8}} .empty{{color:var(--muted)}} footer{{margin-top:32px;color:var(--muted);font-size:12px}}
@media(max-width:760px){{.grid{{grid-template-columns:1fr 1fr}} main{{padding:16px}}}}
@media print{{body{{background:#fff}} main{{max-width:none;padding:0}} .card,.table-wrap{{break-inside:avoid}}}}
</style>
</head>
<body><main>
<header><h1>FinCompiler Management Pack</h1><div class="subtitle">Local-first deterministic month-end evidence · Run {_cell(manifest['run_id'])}</div><span class="badge {ready_class}">{_cell(readiness)}</span></header>
<div class="grid">
<div class="card"><div class="label">Controls complete</div><div class="value">{workflow['completed_tasks']} / {workflow['total_tasks']}</div></div>
<div class="card"><div class="label">Blocking items</div><div class="value">{len(exceptions)}</div></div>
<div class="card"><div class="label">Sales / GL difference</div><div class="value">{_cell(recon['variance'])}</div></div>
<div class="card"><div class="label">Base currency</div><div class="value">{_cell(fx['base_currency'])}</div></div>
</div>
<p class="notice">FinCompiler does not create balancing entries, silently approve uncertain mappings or replace Finance review.</p>
<h2>Month-end action plan</h2>
{_table(tasks, [('question','Question'),('status','Status'),('outcome','Outcome'),('next_action','Next action')])}
<h2>Blocking control items</h2>
{_table(exceptions, [('severity','Severity'),('code','Code'),('message','Message'),('context','Evidence')])}
<h2>Sales vs GL investigator</h2>
<div class="grid">
<div class="card"><div class="label">Sales</div><div class="value">{_cell(recon['sales_total']['value'])}</div></div>
<div class="card"><div class="label">GL revenue</div><div class="value">{_cell(recon['gl_total']['value'])}</div></div>
<div class="card"><div class="label">Difference</div><div class="value">{_cell(recon['variance'])}</div></div>
<div class="card"><div class="label">Status</div><div class="value">{_cell(recon['status'])}</div></div>
</div>
{_table(causes, [('reference','Reference'),('reason','Cause'),('sales','Sales'),('gl','GL'),('difference','Difference')])}
<h2>Budget vs Actual / PVM</h2>
<div class="grid">
<div class="card"><div class="label">Variance</div><div class="value">{_cell(pvm['totals'].get('variance','0.00'))}</div></div>
<div class="card"><div class="label">Volume</div><div class="value">{_cell(pvm['totals'].get('volume','0.00'))}</div></div>
<div class="card"><div class="label">Price</div><div class="value">{_cell(pvm['totals'].get('price','0.00'))}</div></div>
<div class="card"><div class="label">Bridge check</div><div class="value">{_cell(pvm['totals']['bridge_check'])}</div></div>
</div>
{_table(segments, [('customer','Customer'),('sku','SKU'),('actual_revenue','Actual'),('budget_revenue','Budget'),('variance','Variance'),('volume','Volume'),('price','Price'),('mix_residual','Mix / residual')])}
<h2>Applied exchange-rate evidence</h2>
{_table(applications, [('dataset','Dataset'),('record_id','Record'),('source_currency','From'),('target_currency','To'),('effective_date','Rate date'),('rate','Rate'),('method','Method'),('formula','Formula'),('observations','Evidence')])}
<h2>Source files and integrity hashes</h2>
{_table(sources, [('dataset','Dataset'),('file','Local file'),('bytes','Bytes'),('sha256','SHA-256')])}
<footer>Generated by FinCompiler engine {_cell(manifest['engine_version'])}. Created {_cell(manifest['created_at_utc'])}. Full row-level lineage remains in the local SQLite evidence store.</footer>
</main></body></html>"""


def write_management_pack_html(output_dir: str | Path, report: dict[str, Any]) -> Path:
    path = Path(output_dir) / "management_pack.html"
    path.write_text(render_management_pack_html(report), encoding="utf-8")
    return path


def _as_decimal(value: Any) -> Decimal | None:
    if isinstance(value, Decimal):
        return value
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _excel_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def write_management_pack_excel(
    output_dir: str | Path,
    report: dict[str, Any],
    exception_workflow: dict[str, Any] | None = None,
) -> Path:
    """Write a values-only Finance workbook from deterministic report results."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - packaging always includes openpyxl
        raise RuntimeError("Excel export requires openpyxl. Reinstall FinCompiler to restore the workbook dependency.") from exc

    navy = "14233C"
    blue = "225EA8"
    pale_blue = "EAF1F8"
    pale_red = "FDE5E3"
    pale_green = "DFF4E7"
    pale_yellow = "FFF2CC"
    white = "FFFFFF"
    muted = "657086"
    thin_fill = PatternFill("solid", fgColor=pale_blue)
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(color=white, bold=True)
    title_font = Font(color=navy, bold=True, size=18)
    section_font = Font(color=blue, bold=True, size=12)

    workbook = Workbook()
    workbook.remove(workbook.active)
    manifest = report["run_manifest"]
    created = manifest.get("created_at_utc")
    if created:
        timestamp = datetime.fromisoformat(created.replace("Z", "+00:00")).replace(tzinfo=None)
        workbook.properties.created = timestamp
        workbook.properties.modified = timestamp
    workbook.properties.creator = "FinCompiler"
    workbook.properties.title = f"FinCompiler Management Pack {manifest['run_id']}"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False

    def sheet(name: str):
        worksheet = workbook.create_sheet(name)
        worksheet.sheet_view.showGridLines = False
        return worksheet

    def title(worksheet, value: str, subtitle: str = "") -> int:
        worksheet["A1"] = value
        worksheet["A1"].font = title_font
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        row = 2
        if subtitle:
            worksheet["A2"] = subtitle
            worksheet["A2"].font = Font(color=muted, italic=True)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
            row = 3
        return row

    money_format = '#,##0.00;[Red](#,##0.00);-'

    def table(
        worksheet,
        start_row: int,
        rows: list[dict[str, Any]],
        columns: list[tuple[str, str]],
        money_keys: set[str] | None = None,
        status_keys: set[str] | None = None,
        empty_message: str = "No items.",
    ) -> int:
        money_keys = money_keys or set()
        status_keys = status_keys or set()
        for column_index, (_, label) in enumerate(columns, start=1):
            cell = worksheet.cell(start_row, column_index, label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top")
        for row_index, source in enumerate(rows, start=start_row + 1):
            for column_index, (key, _) in enumerate(columns, start=1):
                raw = source.get(key)
                cell = worksheet.cell(row_index, column_index)
                if key in money_keys and (number := _as_decimal(raw)) is not None:
                    cell.value = number
                    cell.number_format = money_format
                else:
                    cell.value = _excel_cell(raw)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if key in status_keys:
                    normalized = str(raw or "").upper()
                    if normalized in {"READY", "PASS", "COMPLETE", "CONFIRMED", "CLEARED_BY_RERUN"}:
                        cell.fill = PatternFill("solid", fgColor=pale_green)
                    elif normalized in {"BLOCKED", "FAIL", "NEEDS_ACTION", "NEEDS_REVIEW"}:
                        cell.fill = PatternFill("solid", fgColor=pale_red)
                    elif normalized in {"OPEN", "INVESTIGATING", "WAITING_FOR_SOURCE_FIX", "READY_TO_RERUN"}:
                        cell.fill = PatternFill("solid", fgColor=pale_yellow)
                    cell.font = Font(bold=True)
        end_row = start_row + max(1, len(rows))
        if rows:
            worksheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(columns))}{end_row}"
        else:
            worksheet.cell(start_row + 1, 1, empty_message)
            worksheet.cell(start_row + 1, 1).font = Font(color=muted, italic=True)
        worksheet.freeze_panes = f"A{start_row + 1}"
        for column_index, (_, label) in enumerate(columns, start=1):
            values = [str(_excel_cell(row.get(columns[column_index - 1][0], ""))) for row in rows[:200]]
            width = min(55, max(12, len(label) + 2, *(len(value) + 2 for value in values)))
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        return end_row + 2

    summary = sheet("Summary")
    row = title(summary, "FinCompiler Management Pack", f"Run {manifest['run_id']} · deterministic values only")
    summary_data = [
        ("Publish readiness", report["output_readiness"]),
        ("Company", report["configuration"].get("company_name")),
        ("Base currency", report["fx"]["base_currency"]),
        ("Controls complete", f"{report['close_workflow']['completed_tasks']} / {report['close_workflow']['total_tasks']}"),
        ("Blocking control items", len(report["exceptions"])),
        ("Sales", report["reconciliation"]["sales_total"]["value"]),
        ("GL revenue", report["reconciliation"]["gl_total"]["value"]),
        ("Sales / GL difference", report["reconciliation"]["variance"]),
        ("Budget variance", report["pvm"]["totals"].get("variance", "0.00")),
        ("PVM bridge check", report["pvm"]["totals"].get("bridge_check", "0.00")),
    ]
    for label, value in summary_data:
        summary.cell(row, 1, label).fill = thin_fill
        summary.cell(row, 1).font = Font(bold=True, color=navy)
        value_cell = summary.cell(row, 2)
        if label in {"Sales", "GL revenue", "Sales / GL difference", "Budget variance", "PVM bridge check"}:
            value_cell.value = _as_decimal(value)
            value_cell.number_format = money_format
        else:
            value_cell.value = value
        row += 1
    summary.cell(row + 1, 1, "Trust boundary").font = section_font
    summary.cell(row + 2, 1, "FinCompiler does not create balancing entries, silently approve uncertain mappings or replace Finance review.")
    summary.merge_cells(start_row=row + 2, start_column=1, end_row=row + 2, end_column=8)
    summary.cell(row + 2, 1).alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 34
    readiness_cell = summary["B3"]
    readiness_cell.fill = PatternFill("solid", fgColor=pale_green if report["output_readiness"] == "READY" else pale_red)
    readiness_cell.font = Font(bold=True)

    mapping_review_count = sum(
        proposal.get("status") == "NEEDS_REVIEW"
        for proposals in report.get("mapping_proposals", {}).values()
        for proposal in proposals
    )
    tolerance = _as_decimal(report["configuration"].get("reconciliation_tolerance", "0.00")) or Decimal("0.00")
    recon_variance = abs(_as_decimal(report["reconciliation"]["variance"]) or Decimal("0.00"))
    bridge_variance = abs(_as_decimal(report["pvm"]["totals"].get("bridge_check", "0.00")) or Decimal("0.00"))
    control_rows = [
        {
            "check": "Unreviewed mappings",
            "actual": mapping_review_count,
            "expected": "0",
            "difference": mapping_review_count,
            "tolerance": "0",
            "status": "PASS" if mapping_review_count == 0 else "FAIL",
            "where_to_fix": "Mappings sheet / Action plan",
        },
        {
            "check": "Missing approved FX rates",
            "actual": report["fx"]["missing_records"],
            "expected": "0",
            "difference": report["fx"]["missing_records"],
            "tolerance": "0",
            "status": "PASS" if report["fx"]["missing_records"] == 0 else "FAIL",
            "where_to_fix": "FX Evidence / approved rate book",
        },
        {
            "check": "Sales vs GL absolute difference",
            "actual": recon_variance,
            "expected": "0.00",
            "difference": recon_variance,
            "tolerance": tolerance,
            "status": "PASS" if recon_variance <= tolerance else "FAIL",
            "where_to_fix": "Sales vs GL",
        },
        {
            "check": "PVM bridge residual",
            "actual": bridge_variance,
            "expected": "0.00",
            "difference": bridge_variance,
            "tolerance": tolerance,
            "status": "PASS" if bridge_variance <= tolerance else "FAIL",
            "where_to_fix": "Budget vs Actual",
        },
        {
            "check": "Blocking exceptions",
            "actual": len(report["exceptions"]),
            "expected": "0",
            "difference": len(report["exceptions"]),
            "tolerance": "0",
            "status": "PASS" if not report["exceptions"] else "FAIL",
            "where_to_fix": "Exceptions",
        },
    ]
    checks = sheet("Control Checks")
    row = title(checks, "Control Checks", f"MODEL STATUS: {'PASS' if report['output_readiness'] == 'READY' else 'FAIL'}")
    table(
        checks,
        row,
        control_rows,
        [
            ("check", "Check"), ("actual", "Actual"), ("expected", "Expected"),
            ("difference", "Difference"), ("tolerance", "Tolerance"),
            ("status", "Status"), ("where_to_fix", "Where to fix"),
        ],
        {"actual", "difference", "tolerance"},
        {"status"},
    )

    actions = sheet("Action Plan")
    row = title(actions, "Month-end Action Plan", "What is complete, what is blocked, and what Finance should do next")
    table(actions, row, report["close_workflow"]["tasks"], [
        ("question", "Question"), ("status", "Status"), ("outcome", "Outcome"), ("next_action", "Next action")
    ], status_keys={"status"})

    exception_state = exception_workflow or {"items": []}
    exceptions = sheet("Exceptions")
    row = title(exceptions, "Exception Workflow", exception_state.get("trust_rule", "Only a deterministic rerun can clear an exception."))
    table(exceptions, row, exception_state.get("items", []), [
        ("exception_id", "Exception ID"), ("active", "Active"), ("status", "Handling status"),
        ("owner", "Owner"), ("code", "Control code"), ("severity", "Severity"),
        ("message", "Message"), ("note", "Working note"), ("evidence_reference", "Evidence reference"),
        ("context", "Source evidence"), ("updated_at_utc", "Updated UTC"),
    ], status_keys={"status"}, empty_message="No current or previously cleared exceptions.")

    reconciliation = sheet("Sales vs GL")
    row = title(reconciliation, "Sales vs GL Investigator", f"Configured variance {report['reconciliation']['variance']}")
    reconciliation.cell(row, 1, "Sales total").font = section_font
    reconciliation.cell(row, 2, _as_decimal(report["reconciliation"]["sales_total"]["value"])).number_format = '#,##0.00'
    reconciliation.cell(row + 1, 1, "GL revenue total").font = section_font
    reconciliation.cell(row + 1, 2, _as_decimal(report["reconciliation"]["gl_total"]["value"])).number_format = '#,##0.00'
    reconciliation.cell(row + 2, 1, "Difference").font = section_font
    reconciliation.cell(row + 2, 2, _as_decimal(report["reconciliation"]["variance"])).number_format = money_format
    row += 4
    row = table(reconciliation, row, report["reconciliation"].get("causes", []), [
        ("reference", "Reference"), ("reason", "Deterministic cause"), ("sales", "Sales"),
        ("gl", "GL"), ("difference", "Difference"), ("component_amount", "Component"),
    ], {"sales", "gl", "difference", "component_amount"}, empty_message="No unexplained reconciliation causes.")
    reconciliation.cell(row, 1, "Matched record groups").font = section_font
    row += 1
    table(reconciliation, row, report["reconciliation"].get("match_groups", []), [
        ("group_id", "Group"), ("method", "Match method"), ("sales_record_ids", "Sales records"),
        ("gl_record_ids", "GL records"), ("sales", "Sales"), ("gl", "GL"), ("residual", "Residual"),
    ], {"sales", "gl", "residual"}, empty_message="No deterministic match groups were found.")

    pvm = sheet("Budget vs Actual")
    row = title(pvm, "Budget vs Actual / PVM", "All driver values were calculated with Decimal; no workbook formulas alter them")
    for label, key in (("Variance", "variance"), ("Volume", "volume"), ("Price", "price"), ("Mix / residual", "mix_residual"), ("Bridge check", "bridge_check")):
        pvm.cell(row, 1, label).font = section_font
        pvm.cell(row, 2, _as_decimal(report["pvm"]["totals"].get(key, "0.00"))).number_format = money_format
        row += 1
    row += 1
    table(pvm, row, report["pvm"].get("segments", []), [
        ("customer", "Customer"), ("sku", "SKU"), ("actual_revenue", "Actual revenue"),
        ("budget_revenue", "Budget revenue"), ("variance", "Variance"), ("volume", "Volume"),
        ("price", "Price"), ("mix_residual", "Mix / residual"),
    ], {"actual_revenue", "budget_revenue", "variance", "volume", "price", "mix_residual"})

    fx_rows = []
    fx_source_rows = []
    for application in report["fx"].get("applications", []):
        observations = application.get("observations", [])
        fx_rows.append(
            {
                **application,
                "evidence_ids": "; ".join(item.get("observation_id", "") for item in observations),
            }
        )
        for observation in observations:
            fx_source_rows.append(
                {
                    "application_record": application.get("record_id"),
                    **observation,
                }
            )

    fx = sheet("FX Evidence")
    row = title(fx, "Applied Exchange-rate Evidence", f"Base currency {report['fx']['base_currency']} · rate type {report['fx']['rate_type']}")
    table(fx, row, fx_rows, [
        ("dataset", "Dataset"), ("record_id", "Record"), ("source_currency", "From"),
        ("source_amount", "Source amount"), ("target_currency", "To"),
        ("converted_amount", "Converted amount"), ("requested_date", "Transaction date"),
        ("effective_date", "Rate date"), ("rate", "Rate"), ("method", "Method"),
        ("formula", "Formula"), ("evidence_ids", "Evidence IDs"),
    ], {"source_amount", "converted_amount"}, empty_message="No currency conversion was required.")

    fx_sources = sheet("FX Sources")
    row = title(fx_sources, "Exchange-rate Source Evidence", "One row per approved rate observation used in a conversion")
    table(fx_sources, row, fx_source_rows, [
        ("application_record", "Applied to record"), ("observation_id", "Evidence ID"),
        ("provider", "Provider"), ("source_url", "Source reference"),
        ("fetched_at", "Fetched UTC"), ("effective_date", "Rate date"),
        ("base_currency", "Base"), ("quote_currency", "Quote"),
        ("rate", "Rate"), ("rate_type", "Rate type"), ("raw_sha256", "Raw SHA-256"),
    ], empty_message="No external rate observations were used.")

    lineage_rows: list[dict[str, Any]] = []

    def add_lineage(output: str, lineage_id: str, calculation: str, dataset: str, record_id: str, source: dict[str, Any]) -> None:
        lineage_rows.append(
            {
                "output": output,
                "lineage_id": lineage_id,
                "calculation": calculation,
                "dataset": dataset,
                "record_id": record_id,
                "file": source.get("file"),
                "sheet": source.get("sheet"),
                "row": source.get("row"),
                "field": source.get("field"),
                "raw_value": source.get("raw_value"),
            }
        )

    sales_total = report["reconciliation"]["sales_total"]
    for item in sales_total.get("input_preview", []):
        add_lineage("Sales total", sales_total.get("lineage_id", ""), sales_total.get("formula", ""), "sales", item.get("record_id", ""), item.get("source", {}))
    gl_total = report["reconciliation"]["gl_total"]
    for item in gl_total.get("input_preview", []):
        derivation = item.get("derivation", {})
        for source in derivation.get("sources", []):
            add_lineage("GL revenue total", gl_total.get("lineage_id", ""), derivation.get("formula", gl_total.get("formula", "")), "gl", item.get("record_id", ""), source)
    for segment in report["pvm"].get("segments", []):
        output = f"PVM · {segment.get('customer', '')} · {segment.get('sku', '')}"
        calculation = (
            "variance = actual revenue - budget revenue; "
            "volume = (actual qty - budget qty) × budget price; "
            "price = actual qty × (actual price - budget price); "
            "mix/residual = variance - volume - price"
        )
        add_lineage(output, segment.get("lineage_id", ""), calculation, "calculation", "", {})
        for dataset, preview_key in (("sales", "actual_input_preview"), ("budget", "budget_input_preview")):
            for item in segment.get(preview_key, []):
                for source_key in ("quantity_source", "revenue_source", "unit_price_source"):
                    source = item.get(source_key)
                    if source:
                        add_lineage(output, segment.get("lineage_id", ""), "", dataset, item.get("record_id", ""), source)

    lineage = sheet("Lineage")
    row = title(lineage, "Source and Calculation Lineage", "Output → calculation → dataset → file → sheet → row → field → raw value")
    table(lineage, row, lineage_rows, [
        ("output", "Output"), ("lineage_id", "Lineage ID"), ("calculation", "Calculation chain"),
        ("dataset", "Dataset"), ("record_id", "Record"), ("file", "Source file"),
        ("sheet", "Sheet"), ("row", "Row"), ("field", "Source field"), ("raw_value", "Raw value"),
    ], empty_message="No preview lineage rows are available; use the local SQLite store for the full trace.")

    audit = sheet("Source Audit")
    row = title(audit, "Source Files and Integrity", "Local source paths and SHA-256 evidence used for this run")
    table(audit, row, manifest.get("sources", []), [
        ("dataset", "Dataset"), ("file", "Local file"), ("bytes", "Bytes"), ("sha256", "SHA-256")
    ])

    mappings = sheet("Mappings")
    row = title(mappings, "Canonical Mapping Decisions", "Low-confidence mappings remain explicit and cannot be silently confirmed")
    mapping_rows = []
    for dataset, proposals in report.get("mapping_proposals", {}).items():
        mapping_rows.extend({"dataset": dataset, **proposal} for proposal in proposals)
    table(mappings, row, mapping_rows, [
        ("dataset", "Dataset"), ("source_field", "Source field"), ("canonical_field", "Canonical field"),
        ("status", "Status"), ("confidence", "Confidence"), ("reason", "Reason"),
    ], status_keys={"status"})

    path = Path(output_dir) / "management_pack.xlsx"
    workbook.save(path)
    return path

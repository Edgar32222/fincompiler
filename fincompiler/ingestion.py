from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator

from .models import SourceRef


SUPPORTED_INPUT_EXTENSIONS = (".csv", ".xlsx", ".xlsm")


def discover_dataset_files(input_dir: str | Path, datasets: tuple[str, ...] = ("sales", "gl", "budget")) -> dict[str, Path]:
    """Find one explicitly named CSV/Excel file for every required dataset.

    Discovery is deliberately conservative: FinCompiler accepts case-insensitive
    ``sales``, ``gl`` and ``budget`` stems, but never guesses between multiple
    candidate files.
    """
    directory = Path(input_dir)
    if not directory.exists():
        raise FileNotFoundError(f"Input folder not found: {directory}")
    if not directory.is_dir():
        raise NotADirectoryError(f"Input path must be a folder: {directory}")
    files = [path for path in directory.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS]
    discovered: dict[str, Path] = {}
    for dataset in datasets:
        matches = [path for path in files if path.stem.strip().lower() == dataset]
        if not matches:
            accepted = ", ".join(f"{dataset}{suffix}" for suffix in SUPPORTED_INPUT_EXTENSIONS)
            raise FileNotFoundError(f"Missing {dataset.title()} file in {directory}. Add one of: {accepted}")
        if len(matches) > 1:
            names = ", ".join(sorted(path.name for path in matches))
            raise ValueError(f"Multiple {dataset.title()} files found ({names}). Keep exactly one to avoid choosing silently.")
        discovered[dataset] = matches[0]
    return discovered


def _csv_rows(path: Path) -> Iterator[tuple[str, int, dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row_no, row in enumerate(csv.DictReader(handle), start=2):
            yield "CSV", row_no, {str(k).strip(): (v or "").strip() for k, v in row.items()}


def _xlsx_rows(path: Path) -> Iterator[tuple[str, int, dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel input requires the optional 'excel' dependency") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            first_row = next(rows)
        except StopIteration:
            continue
        headers = [str(value).strip() if value is not None else "" for value in first_row]
        if not any(headers):
            continue
        for row_no, values in enumerate(rows, start=2):
            if not any(value not in {None, ""} for value in values):
                continue
            yield sheet.title, row_no, {
                header: "" if value is None else str(value)
                for header, value in zip(headers, values)
                if header
            }


def read_tabular(path: str | Path) -> list[tuple[dict[str, str], dict[str, SourceRef]]]:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Input file not found: {source}")
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED_INPUT_EXTENSIONS:
        raise ValueError(f"Unsupported input type '{suffix}'. Use CSV or XLSX.")
    iterator = _csv_rows(source) if suffix == ".csv" else _xlsx_rows(source)
    output = []
    for sheet, row_no, row in iterator:
        refs = {
            field: SourceRef(str(source.resolve()), sheet, row_no, field, value)
            for field, value in row.items()
        }
        output.append((row, refs))
    if not output:
        raise ValueError(f"No data rows found in {source.name}. Include one header row and at least one data row.")
    return output

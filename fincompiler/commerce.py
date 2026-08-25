from __future__ import annotations

import json
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path

from . import __version__
from .config import CommerceConfig
from .fx import RateBook, quantize_currency, validate_currency
from .ingestion import SUPPORTED_INPUT_EXTENSIONS, read_tabular
from .mapping import MappingMemory, apply_mapping
from .models import CanonicalRecord, ExceptionItem, money
from .normalization import normalize_records
from .profiles import detect_profile
from .run_manifest import build_manifest
from .run_state import assert_writable, initialize_draft


COMMERCE_DATASETS = ("amazon_settlements", "shopify_orders", "shopify_payouts", "bank", "sku_costs")
REQUIRED_FIELDS = {
    "amazon_settlements": {"settlement_id", "payout_date", "settlement_total", "currency", "amount_type", "amount_description", "amount"},
    "shopify_orders": {"order_id", "date", "currency", "gross_sales", "quantity", "sku", "unit_price"},
    "shopify_payouts": {"date", "transaction_type", "payout_status", "payout_date", "net_amount"},
    "bank": {"bank_transaction_id", "date", "bank_reference", "amount", "currency"},
    "sku_costs": {"sku", "effective_date", "unit_purchase_cost", "currency"},
}


def discover_commerce_files(input_dir: str | Path) -> dict[str, Path]:
    directory = Path(input_dir)
    if not directory.is_dir():
        raise FileNotFoundError(f"Commerce input folder not found: {directory}")
    files = [item for item in directory.iterdir() if item.is_file() and item.suffix.lower() in SUPPORTED_INPUT_EXTENSIONS]
    found: dict[str, Path] = {}
    for dataset in COMMERCE_DATASETS:
        matches = [item for item in files if item.stem.strip().lower() == dataset]
        if len(matches) > 1:
            raise ValueError(f"Multiple {dataset} files found; keep exactly one so FinCompiler does not choose silently.")
        if matches:
            found[dataset] = matches[0]
    if "bank" not in found:
        raise FileNotFoundError("Missing bank.csv/xlsx. A bank statement is required to verify cash received.")
    if not ({"amazon_settlements", "shopify_payouts"} & found.keys()):
        raise FileNotFoundError("Add amazon_settlements or shopify_payouts so platform cash can be compared with the bank.")
    return found


def _source(ref) -> dict:
    return {"file": ref.file, "sheet": ref.sheet, "row": ref.row, "field": ref.field, "raw_value": ref.raw_value}


def _required_field_exceptions(dataset: str, records: list[CanonicalRecord]) -> list[ExceptionItem]:
    if not records:
        return []
    available = set().union(*(record.values.keys() for record in records))
    missing = sorted(REQUIRED_FIELDS[dataset] - available)
    if not missing:
        return []
    return [ExceptionItem("REQUIRED_FIELD_MISSING", "BLOCKING", "A required commerce field is absent after mapping", {"dataset": dataset, "fields": missing})]


def _convert_fields(
    records: list[CanonicalRecord],
    dataset: str,
    fields: tuple[str, ...],
    config: CommerceConfig,
    rate_book: RateBook | None,
) -> tuple[list[ExceptionItem], list[dict]]:
    exceptions: list[ExceptionItem] = []
    evidence: list[dict] = []
    for record in records:
        try:
            source_currency = validate_currency(record.values.get("currency") or config.base_currency)
        except ValueError as exc:
            exceptions.append(ExceptionItem("TYPE_VALIDATION_FAILED", "BLOCKING", "Currency is not a supported ISO 4217 code", {"dataset": dataset, "record_id": record.record_id, "reason": str(exc)}))
            continue
        record.values["transaction_currency"] = source_currency
        record.values["currency"] = source_currency
        if source_currency == config.base_currency:
            record.values["basis_currency"] = config.base_currency
            continue
        requested_raw = record.values.get("date") or record.values.get("payout_date") or record.values.get("effective_date")
        try:
            requested = date.fromisoformat(str(requested_raw)[:10])
            if rate_book is None:
                raise LookupError("no approved local rate book configured")
            match = rate_book.find(source_currency, config.base_currency, requested, config.fx_policy)
        except (LookupError, ValueError) as exc:
            exceptions.append(ExceptionItem("FX_RATE_REQUIRED", "BLOCKING", "A commerce amount has no approved conversion basis", {"dataset": dataset, "record_id": record.record_id, "currency": source_currency, "base_currency": config.base_currency, "date": str(requested_raw), "reason": str(exc)}))
            continue
        for field in fields:
            if field not in record.values or record.values[field] in {None, ""}:
                continue
            original = money(record.values[field])
            converted = quantize_currency(original * match.rate, config.base_currency)
            record.values[f"transaction_{field}"] = original
            record.values[field] = converted
            record.derivations[field] = {"formula": f"transaction_{field} * applied_exchange_rate", "sources": [_source(record.lineage[field])] if field in record.lineage else [], "rate_evidence": match.jsonable()}
            evidence.append({"dataset": dataset, "record_id": record.record_id, "field": field, "source_amount": str(original), "converted_amount": str(converted), **match.jsonable()})
        record.values["currency"] = config.base_currency
        record.values["basis_currency"] = config.base_currency
        record.values["applied_exchange_rate"] = match.rate
    return exceptions, evidence


def _platform_payouts(records: dict[str, list[CanonicalRecord]], tolerance: Decimal) -> tuple[list[dict], list[ExceptionItem]]:
    payouts: list[dict] = []
    exceptions: list[ExceptionItem] = []
    amazon_groups: dict[str, list[CanonicalRecord]] = defaultdict(list)
    for record in records.get("amazon_settlements", []):
        amazon_groups[str(record.values.get("settlement_id", ""))].append(record)
    for settlement_id, group in sorted(amazon_groups.items()):
        totals = {money(item.values.get("settlement_total")) for item in group}
        if len(totals) != 1:
            exceptions.append(ExceptionItem("INCONSISTENT_SETTLEMENT_TOTAL", "BLOCKING", "Amazon rows disagree on total-amount", {"settlement_id": settlement_id, "totals": sorted(map(str, totals)), "record_ids": [item.record_id for item in group]}))
            continue
        settlement_total = next(iter(totals))
        activity_total = sum((money(item.values.get("amount")) for item in group), Decimal("0")).quantize(Decimal("0.01"))
        activity_difference = (settlement_total - activity_total).quantize(Decimal("0.01"))
        if abs(activity_difference) > tolerance:
            exceptions.append(ExceptionItem("SETTLEMENT_ACTIVITY_INCOMPLETE", "BLOCKING", "Amazon settlement activity does not add back to total-amount", {"settlement_id": settlement_id, "settlement_total": str(settlement_total), "activity_total": str(activity_total), "difference": str(activity_difference), "record_ids": [item.record_id for item in group]}))
        payouts.append({"payout_id": f"amazon:{settlement_id}", "platform": "Amazon", "expected_amount": str(settlement_total), "activity_total": str(activity_total), "activity_difference": str(activity_difference), "payout_date": str(group[0].values.get("payout_date", "")), "currency": str(group[0].values.get("currency", "")), "bank_reference": settlement_id, "source_record_ids": [item.record_id for item in group]})

    shopify_groups: dict[tuple[str, str, str], list[CanonicalRecord]] = defaultdict(list)
    for record in records.get("shopify_payouts", []):
        if str(record.values.get("payout_status", "")).strip().lower() != "paid":
            continue
        key = (str(record.values.get("payout_id", "")), str(record.values.get("payout_date", "")), str(record.values.get("currency", "")))
        shopify_groups[key].append(record)
    for (source_payout_id, payout_date, currency), group in sorted(shopify_groups.items()):
        expected = sum((money(item.values.get("net_amount")) for item in group), Decimal("0")).quantize(Decimal("0.01"))
        group_id = source_payout_id or f"{payout_date}:{currency}"
        refs = {str(item.values.get("bank_reference", "")).strip() for item in group} - {""}
        payouts.append({"payout_id": f"shopify:{group_id}", "platform": "Shopify", "expected_amount": str(expected), "payout_date": payout_date, "currency": currency, "bank_reference": next(iter(refs)) if len(refs) == 1 else source_payout_id, "source_record_ids": [item.record_id for item in group]})
    return payouts, exceptions


def _match_bank(payouts: list[dict], bank: list[CanonicalRecord], config: CommerceConfig) -> tuple[list[dict], list[ExceptionItem]]:
    results: list[dict] = []
    exceptions: list[ExceptionItem] = []
    used: set[str] = set()
    for payout in payouts:
        expected = money(payout["expected_amount"])
        payout_date = date.fromisoformat(str(payout["payout_date"])[:10])
        eligible = [item for item in bank if item.record_id not in used and item.values.get("date") and abs((date.fromisoformat(str(item.values["date"])[:10]) - payout_date).days) <= config.bank_match_window_days]
        reference = str(payout.get("bank_reference", "")).strip().lower()
        referenced = [item for item in eligible if reference and reference in str(item.values.get("bank_reference", "")).lower()]
        exact_amount = [item for item in eligible if abs(money(item.values.get("amount")) - expected) <= config.reconciliation_tolerance]
        candidate = referenced[0] if len(referenced) == 1 else (exact_amount[0] if len(exact_amount) == 1 else None)
        method = "EXACT_REFERENCE" if len(referenced) == 1 else ("UNIQUE_AMOUNT_DATE_WINDOW" if len(exact_amount) == 1 else "UNMATCHED")
        if candidate is None and len(eligible) == 1:
            candidate, method = eligible[0], "UNIQUE_DATE_WINDOW"
        if candidate is None:
            code = "AMBIGUOUS_BANK_MATCH" if eligible else "BANK_RECEIPT_MISSING"
            exceptions.append(ExceptionItem(code, "BLOCKING", "Platform payout could not be matched safely to one bank receipt", {"payout_id": payout["payout_id"], "expected_amount": str(expected), "eligible_bank_record_ids": [item.record_id for item in eligible]}))
            results.append({**payout, "status": "NEEDS_REVIEW", "method": method, "bank_amount": "0.00", "difference": str(expected), "bank_record_id": None, "reason": code})
            continue
        used.add(candidate.record_id)
        actual = money(candidate.values.get("amount"))
        difference = (expected - actual).quantize(Decimal("0.01"))
        status = "PASS" if abs(difference) <= config.reconciliation_tolerance else "NEEDS_REVIEW"
        reason = "MATCHED" if status == "PASS" else "BANK_OR_PROCESSOR_DEDUCTION"
        if status != "PASS":
            exceptions.append(ExceptionItem("PAYOUT_BANK_VARIANCE", "BLOCKING", "A specific payout and bank receipt differ", {"payout_id": payout["payout_id"], "bank_record_id": candidate.record_id, "expected_amount": str(expected), "bank_amount": str(actual), "difference": str(difference), "reason": reason}))
        results.append({**payout, "status": status, "method": method, "bank_amount": str(actual), "difference": str(difference), "bank_record_id": candidate.record_id, "reason": reason})
    for record in bank:
        if record.record_id not in used and money(record.values.get("amount")) > 0:
            exceptions.append(ExceptionItem("UNMATCHED_BANK_RECEIPT", "HIGH", "A positive bank receipt was not linked to a platform payout", {"bank_record_id": record.record_id, "amount": str(money(record.values.get("amount"))), "reference": str(record.values.get("bank_reference", ""))}))
    return results, exceptions


def _cost_map(cost_records: list[CanonicalRecord]) -> dict[str, list[CanonicalRecord]]:
    result: dict[str, list[CanonicalRecord]] = defaultdict(list)
    for record in cost_records:
        result[str(record.values.get("sku", "")).strip()].append(record)
    for values in result.values():
        values.sort(key=lambda item: str(item.values.get("effective_date", "")))
    return result


def _unit_cost(sku: str, activity_date: str, costs: dict[str, list[CanonicalRecord]]) -> Decimal | None:
    eligible = [item for item in costs.get(sku, []) if str(item.values.get("effective_date", "")) <= activity_date]
    if not eligible:
        return None
    record = eligible[-1]
    return sum((money(record.values.get(field)) for field in ("unit_purchase_cost", "unit_freight_cost", "unit_duty_cost", "other_unit_cost")), Decimal("0"))


def _profitability(records: dict[str, list[CanonicalRecord]]) -> tuple[list[dict], list[ExceptionItem]]:
    metrics: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(lambda: {"revenue": Decimal("0"), "platform_fees": Decimal("0"), "quantity": Decimal("0"), "cogs": Decimal("0")})
    exceptions: list[ExceptionItem] = []
    costs = _cost_map(records.get("sku_costs", []))

    payout_by_order: dict[str, list[CanonicalRecord]] = defaultdict(list)
    for item in records.get("shopify_payouts", []):
        order_id = str(item.values.get("order_id", "")).strip()
        transaction_type = str(item.values.get("transaction_type", "")).strip().lower()
        if transaction_type not in {"charge", "payment", "sale", "refund"}:
            exceptions.append(ExceptionItem("UNCLASSIFIED_SHOPIFY_ACTIVITY", "HIGH", "A Shopify payout activity is preserved but excluded from SKU profit until classified", {"record_id": item.record_id, "transaction_type": item.values.get("transaction_type"), "gross_amount": str(money(item.values.get("gross_amount"))), "fee_amount": str(money(item.values.get("fee_amount"))), "net_amount": str(money(item.values.get("net_amount")))}))
        if not order_id and any(money(item.values.get(field)) for field in ("gross_amount", "fee_amount", "net_amount")):
            exceptions.append(ExceptionItem("SHOPIFY_ACTIVITY_NOT_ATTRIBUTED", "HIGH", "A Shopify payout activity affects cash but has no order key for safe SKU allocation", {"record_id": item.record_id, "transaction_type": item.values.get("transaction_type"), "net_amount": str(money(item.values.get("net_amount")))}))
        if order_id:
            payout_by_order[order_id].append(item)
    order_groups: dict[str, list[CanonicalRecord]] = defaultdict(list)
    for item in records.get("shopify_orders", []):
        order_groups[str(item.values.get("order_id", ""))].append(item)
    for order_id, lines in sorted(order_groups.items()):
        totals = {money(item.values.get("gross_sales")) for item in lines}
        taxes = {money(item.values.get("tax_amount")) for item in lines}
        if len(totals) != 1 or len(taxes) > 1:
            exceptions.append(ExceptionItem("INCONSISTENT_ORDER_TOTAL", "BLOCKING", "Shopify line rows disagree on order-level Total or Taxes", {"order_id": order_id, "record_ids": [item.record_id for item in lines]}))
            continue
        distributable = next(iter(totals)) - (next(iter(taxes)) if taxes else Decimal("0"))
        weights = [money(item.values.get("unit_price")) * money(item.values.get("quantity")) for item in lines]
        denominator = sum(weights, Decimal("0"))
        if denominator == 0:
            exceptions.append(ExceptionItem("ORDER_ALLOCATION_BLOCKED", "BLOCKING", "Shopify order total cannot be allocated because line value is zero", {"order_id": order_id}))
            continue
        order_fee = sum((abs(money(item.values.get("fee_amount"))) for item in payout_by_order.get(order_id, [])), Decimal("0"))
        refund_adjustment = sum((money(item.values.get("gross_amount")) for item in payout_by_order.get(order_id, []) if "refund" in str(item.values.get("transaction_type", "")).lower()), Decimal("0"))
        allocated_revenue = Decimal("0")
        allocated_fee = Decimal("0")
        for index, (line, weight) in enumerate(zip(lines, weights)):
            share = weight / denominator
            revenue = (distributable + refund_adjustment) * share
            fee = order_fee * share
            if index == len(lines) - 1:
                revenue = distributable + refund_adjustment - allocated_revenue
                fee = order_fee - allocated_fee
            revenue, fee = revenue.quantize(Decimal("0.01")), fee.quantize(Decimal("0.01"))
            allocated_revenue += revenue
            allocated_fee += fee
            sku = str(line.values.get("sku", "")).strip() or "(missing SKU)"
            quantity = money(line.values.get("quantity"))
            unit_cost = _unit_cost(sku, str(line.values.get("date", "")), costs)
            if unit_cost is None:
                exceptions.append(ExceptionItem("SKU_COST_REQUIRED", "BLOCKING", "True SKU profit is blocked until an effective landed cost is supplied", {"platform": "Shopify", "sku": sku, "order_id": order_id, "record_id": line.record_id}))
                cogs = Decimal("0")
            else:
                cogs = (unit_cost * quantity).quantize(Decimal("0.01"))
            metric = metrics[("Shopify", sku)]
            metric["revenue"] += revenue
            metric["platform_fees"] += fee
            metric["quantity"] += quantity
            metric["cogs"] += cogs

    amazon_seen_qty: set[tuple[str, str]] = set()
    known_revenue = {"principal", "shipping", "giftwrap"}
    known_tax = {"tax", "shippingtax", "giftwraptax"}
    for item in records.get("amazon_settlements", []):
        amount_type = str(item.values.get("amount_type", "")).strip().lower()
        description = str(item.values.get("amount_description", "")).strip().replace(" ", "").lower()
        sku = str(item.values.get("sku", "")).strip() or "(not attributable)"
        amount = money(item.values.get("amount"))
        metric = metrics[("Amazon", sku)]
        if amount_type == "itemprice" and description in known_revenue:
            metric["revenue"] += amount
        elif amount_type == "itemprice" and description in known_tax:
            pass
        elif amount_type in {"itemfees", "servicefees"}:
            metric["platform_fees"] += -amount
            if sku == "(not attributable)":
                exceptions.append(ExceptionItem("FEE_NOT_ATTRIBUTABLE_TO_SKU", "HIGH", "A platform fee affects store profit but has no SKU key for safe allocation", {"record_id": item.record_id, "amount": str(amount), "amount_description": item.values.get("amount_description")}))
        elif amount_type in {"promotion", "othertransaction"}:
            metric["revenue"] += amount
        else:
            exceptions.append(ExceptionItem("UNCLASSIFIED_AMAZON_ACTIVITY", "HIGH", "An Amazon settlement activity is preserved but excluded from SKU profit until classified", {"record_id": item.record_id, "amount_type": item.values.get("amount_type"), "amount_description": item.values.get("amount_description"), "amount": str(amount)}))
        qty_key = (str(item.values.get("order_id", "")), sku)
        if qty_key not in amazon_seen_qty and str(item.values.get("transaction_type", "")).strip().lower() == "order":
            amazon_seen_qty.add(qty_key)
            quantity = money(item.values.get("quantity"))
            metric["quantity"] += quantity
            unit_cost = _unit_cost(sku, str(item.values.get("date", "")), costs)
            if unit_cost is None and sku != "(not attributable)":
                exceptions.append(ExceptionItem("SKU_COST_REQUIRED", "BLOCKING", "True SKU profit is blocked until an effective landed cost is supplied", {"platform": "Amazon", "sku": sku, "order_id": item.values.get("order_id"), "record_id": item.record_id}))
            elif unit_cost is not None:
                metric["cogs"] += (unit_cost * quantity).quantize(Decimal("0.01"))

    output = []
    for (platform, sku), metric in sorted(metrics.items()):
        profit = (metric["revenue"] - metric["platform_fees"] - metric["cogs"]).quantize(Decimal("0.01"))
        output.append({"platform": platform, "sku": sku, **{key: str(value.quantize(Decimal("0.01"))) for key, value in metric.items()}, "profit": str(profit), "formula": "revenue - platform_fees - landed_cogs"})
    return output, exceptions


def _write_excel(path: Path, report: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    numeric_fields = {"expected_amount", "bank_amount", "difference", "quantity", "revenue", "platform_fees", "cogs", "profit"}

    def excel_value(field: str, value):
        if field in numeric_fields and value not in {None, ""}:
            return float(Decimal(str(value)))
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False)
        return value

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["FinCompiler Cross-border Cash & Profit Pack", None])
    summary.append(["Run ID", report["run_manifest"]["run_id"]])
    summary.append(["Output readiness", report["output_readiness"]])
    summary.append(["Base currency", report["configuration"]["base_currency"]])
    summary.append(["Payouts checked", len(report["payout_reconciliation"])])
    summary.append(["Payouts needing review", sum(item["status"] != "PASS" for item in report["payout_reconciliation"])])
    summary.merge_cells("A1:B1")
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="17365D")
    summary["A1"].alignment = Alignment(horizontal="left")
    summary.column_dimensions["A"].width = 48
    summary.column_dimensions["B"].width = 24
    for title, rows, fields in (
        ("Payout Reconciliation", report["payout_reconciliation"], ["platform", "payout_id", "payout_date", "expected_amount", "bank_amount", "difference", "status", "method", "reason", "bank_record_id"]),
        ("SKU Profitability", report["sku_profitability"], ["platform", "sku", "quantity", "revenue", "platform_fees", "cogs", "profit", "formula"]),
        ("Exceptions", report["exceptions"], ["code", "severity", "message", "context"]),
    ):
        sheet = workbook.create_sheet(title)
        sheet.append(fields)
        for row in rows:
            sheet.append([excel_value(field, row.get(field)) for field in fields])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, field in enumerate(fields, start=1):
            if field in numeric_fields:
                for row_index in range(2, sheet.max_row + 1):
                    sheet.cell(row_index, column_index).number_format = '#,##0.00;[Red](#,##0.00);-'
            if field == "status":
                for row_index in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row_index, column_index)
                    cell.fill = PatternFill("solid", fgColor="E2F0D9" if cell.value == "PASS" else "FCE4D6")
    lineage = workbook.create_sheet("Source Lineage")
    lineage.append(["dataset", "record_id", "canonical_field", "file", "sheet", "row", "source_field", "raw_value"])
    for dataset, rows in report["source_lineage"].items():
        for row in rows:
            lineage.append([dataset, row["record_id"], row["canonical_field"], row["file"], row["sheet"], row["row"], row["field"], row["raw_value"]])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A2"
        for column_index, column in enumerate(sheet.columns, start=1):
            width = min(max((len(str(cell.value or "")) for cell in column), default=8) + 2, 48)
            letter = get_column_letter(column_index)
            sheet.column_dimensions[letter].width = max(sheet.column_dimensions[letter].width or 0, width)
    workbook.save(path)


def compile_commerce_pack(input_dir: str | Path, output_dir: str | Path, memory_path: str | Path, config_path: str | Path | None = None) -> dict:
    input_dir, output_dir = Path(input_dir), Path(output_dir)
    assert_writable(output_dir)
    files = discover_commerce_files(input_dir)
    discovered_config = Path(config_path) if config_path else input_dir / "commerce_config.json"
    config = CommerceConfig.load(discovered_config if discovered_config.exists() else None)
    rate_path = None
    if config.fx_rate_book:
        candidate = Path(config.fx_rate_book)
        rate_path = candidate if candidate.is_absolute() else discovered_config.parent / candidate
        if not rate_path.exists():
            raise FileNotFoundError(f"Configured FX rate book not found: {rate_path}")
    manifest = build_manifest(input_dir, config.jsonable(), discovered_config if discovered_config.exists() else None, [("fx_rate_book", rate_path)] if rate_path else None, files)
    memory = MappingMemory(memory_path)
    records: dict[str, list[CanonicalRecord]] = {}
    proposals_dump: dict[str, list[dict]] = {}
    profiles: dict[str, str] = {}
    exceptions: list[ExceptionItem] = []
    for dataset, path in files.items():
        rows = read_tabular(path)
        fields = list(rows[0][0])
        if any(list(row.keys()) != fields for row, _ in rows):
            raise ValueError(f"{path.name} contains multiple header layouts; split it into one table before running.")
        profile = detect_profile(dataset, fields)
        profiles[dataset] = profile.name if profile else "generic"
        proposals, drift = memory.propose(dataset, fields, profile.aliases if profile else None, profile.name if profile else "generic", profile.ignored_fields if profile else None)
        mapped, mapping_issues = apply_mapping(dataset, rows, proposals)
        normalized, type_issues = normalize_records(mapped)
        records[dataset] = normalized
        exceptions.extend(drift + mapping_issues + type_issues + _required_field_exceptions(dataset, normalized))
        proposals_dump[dataset] = [{**item.__dict__, "confidence": str(item.confidence)} for item in proposals]
    rate_book = RateBook.load(rate_path) if rate_path else None
    fx_evidence: list[dict] = []
    fields_by_dataset = {
        "amazon_settlements": ("settlement_total", "amount"),
        "shopify_orders": ("gross_sales", "discount_amount", "shipping_income", "tax_amount", "unit_price"),
        "shopify_payouts": ("gross_amount", "fee_amount", "net_amount"),
        "bank": ("amount",),
        "sku_costs": ("unit_purchase_cost", "unit_freight_cost", "unit_duty_cost", "other_unit_cost"),
    }
    for dataset, dataset_records in records.items():
        fx_issues, applied = _convert_fields(dataset_records, dataset, fields_by_dataset[dataset], config, rate_book)
        exceptions.extend(fx_issues)
        fx_evidence.extend(applied)
    if "shopify_payouts" in records and "shopify_orders" not in records:
        exceptions.append(ExceptionItem("SHOPIFY_ORDERS_REQUIRED_FOR_PROFIT", "BLOCKING", "Shopify payout cash can be checked, but true SKU profit requires the matching order export", {}))
    payouts, payout_issues = _platform_payouts(records, config.reconciliation_tolerance)
    reconciliation, bank_issues = _match_bank(payouts, records.get("bank", []), config)
    profitability, profit_issues = _profitability(records)
    exceptions.extend(payout_issues + bank_issues + profit_issues)
    exception_dump = [item.__dict__ for item in exceptions]
    source_lineage: dict[str, list[dict]] = {}
    for dataset, dataset_records in records.items():
        source_lineage[dataset] = [{"record_id": record.record_id, "canonical_field": field, **_source(ref)} for record in dataset_records for field, ref in record.lineage.items()]
    report = {
        "version": __version__,
        "workflow": "cross_border_cash_and_profit",
        "run_manifest": manifest,
        "configuration": config.jsonable(),
        "source_profiles": profiles,
        "mapping_proposals": proposals_dump,
        "fx": {"base_currency": config.base_currency, "rate_book": str(rate_path.resolve()) if rate_path else None, "applications": fx_evidence},
        "payout_reconciliation": reconciliation,
        "sku_profitability": profitability,
        "exceptions": exception_dump,
        "source_lineage": source_lineage,
        "output_readiness": "BLOCKED" if exceptions or any(item["status"] != "PASS" for item in reconciliation) else "READY",
        "calculation_policy": "All amounts are calculated with Decimal by the deterministic engine. LLM output, if added, may explain but never calculate or adjust values.",
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "commerce_pack.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    (output_dir / "run_manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
    _write_excel(output_dir / "commerce_pack.xlsx", report)
    initialize_draft(output_dir, manifest["run_id"])
    return report
